import openpyxl
import os
from PythonSupports.DBconnection import queryForGettingLineNumbers
from pathlib import Path

if Path("../Resources/support.xlsx").exists():
    path = os.path.abspath("../Resources/support.xlsx")
else:
    path = os.path.abspath("Resources/support.xlsx")

workbook = openpyxl.load_workbook(path)


def readCell(sheetName, r, c):
    sheet = workbook[sheetName]
    return sheet.cell(r, int(c)).value


def writeCell(sheetName, r, c,value):
    colName = ['A','B','C','D','E']
    sheet = workbook[sheetName]
    sheet.cell(row=int(r), column=int(c), value=value)
    sheet.column_dimensions[colName[int(c)-1]].width = len(value)*2
    workbook.save(path)

def getLastRowPlusOne(sheetName):
    sheet = workbook[sheetName]
    return sheet.max_row + 1

def getLastRow(sheetName):
    sheet = workbook[sheetName]
    return sheet.max_row


def createAndDeleteSheet():
    #creating a Temporary sheet for ignoring the error
    workbook.create_sheet('Additional')
    workbook.save(path)

    sheetNames = ['line numbers', 'Order#', 'Accounts', 'API Order#','Fin Trans R3']

    #deleting all sheets except the temporary sheet
    for sheet in workbook.get_sheet_names():
        if sheet == 'Additional':
            pass
        else:
            workbook.remove_sheet(workbook.get_sheet_by_name(sheet))
            workbook.save(path)

    #creating the new sheets
    for sheet in sheetNames:
        workbook.create_sheet(sheet)
        workbook.save(path)

    #deleting the temporary sheet
    workbook.remove_sheet(workbook.get_sheet_by_name('Additional'))
    workbook.save(path)

def writeNewDataToSheet():
    #LineNumber
    r = getLastRow('line numbers')
    writeCell('line numbers',r,1,'Line#')
    writeCell('line numbers', r,2, 'Qty')
    #OrderNumber
    r = getLastRow('Order#')
    writeCell('Order#', r, 1, 'Order No')
    writeCell('Order#', r, 2, 'Master Order No')
    #Accounts
    r = getLastRow('Accounts')
    writeCell('Accounts', r, 1, 'Email')
    writeCell('Accounts', r, 2, 'Type')
    writeCell('Accounts', r, 3, 'Bee No.')
    writeCell('Accounts', r, 4, 'Notes')

    # Entering product numbers from DB in sheet line numbers
    products = queryForGettingLineNumbers('202203')
    for prod in products:
        r = getLastRowPlusOne('line numbers')
        writeCell('line numbers',r,1,prod)
        writeCell('line numbers', r, 2,'2')