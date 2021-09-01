import openpyxl
import os
from openpyxl import Workbook

path = os.path.abspath("Resources/support.xlsx")

workbook = openpyxl.load_workbook(path)


def readCell(sheetName, r, c):
    sheet = workbook[sheetName]
    return sheet.cell(r, int(c)).value


def writeCell(sheetName, r, c,value):
    colName = ['A','B','C','D','E']
    sheet = workbook[sheetName]
    sheet.cell(row=int(r), column=int(c), value=value)
    sheet.column_dimensions[colName[int(c)-1]].width = len(value)*2
    workbook.save(path)

def getLastRowPlusOne(sheetName):
    sheet = workbook[sheetName]
    return sheet.max_row + 1

def getLastRow(sheetName):
    sheet = workbook[sheetName]
    return sheet.max_row
